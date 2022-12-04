"""Модуль, отвечающий за создание статистических отчётов"""
import math
import numpy as np
import matplotlib.pyplot as pyplot
from jinja2 import Environment, FileSystemLoader
import pdfkit
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from utils import Utils


class Stats:
    """
    Класс для представления и вывода статистических данных

    Attributes:
        profession (str): Профессия, по которой требуется статистика
        total_vacancies (int): Общее число вакансий
        year_salary_dynamics (dict[int, int]): Динамика уровня зарплат по годам
        num_of_vacancies_per_year (dict[int, int]): Динамика количества вакансий по годам
        year_salary_dynamics_for_prof (dict[int, int]): Динамика уровня зарплат по годам для выбранной профессии
        num_of_vacancies_per_year_for_prof (dict[int, int]): Динамика количества вакансий по годам для выбранной профессии
        salary_levels_of_areas (dict[str, int]): Уровень зарплат по городам (в порядке убывания)
        vacancy_fractions_of_areas (dict[str, float]): Доля вакансий по городам (в порядке убывания)
    """
    def __init__(self, data):
        """
        Инициализация объекта класса. Получение статистических данных из DataSet

        Args:
            data (DataSet): Объект DataSet, содержащий данные о вакансиях
        """
        self.profession = input('Введите название профессии: ')
        self.total_vacancies = data.length()
        self.year_salary_dynamics = self.get_year_salary_dynamics(data)
        self.num_of_vacancies_per_year = self.get_num_of_vacancies_per_year(data)
        self.year_salary_dynamics_for_prof = self.get_year_salary_dynamics(data, self.profession)
        self.num_of_vacancies_per_year_for_prof = self.get_num_of_vacancies_per_year(data, self.profession)
        self.salary_levels_of_areas = self.get_salary_levels_of_areas(data)
        self.vacancy_fractions_of_areas = self.get_vacancy_fractions_of_areas(data)

    def get_year_salary_dynamics(self, data, profession=''):
        """
        Возвращает словарь, состоящий из пар 'год - средняя зарплата' для всего DataSet или для указанной профессии

        Args:
            data (DataSet): Объект DataSet, содержащий список вакансий
            profession (str): Название профессии, для которой нужно составить статистику. При отсутствии этого аргумента возвращается статистика по всем вакансиям

        Return:
            dict[int, int]: словарь, состоящий из пар 'год - средняя зарплата'
        """
        vacancies = data.vacancies_objects
        vacancies_of_years = Utils.split_list(vacancies, lambda x: Utils.get_year(x.published_at))
        mean_salaries_of_years = {}
        for year in vacancies_of_years:
            sum = 0
            count = 0
            for vacancy in vacancies_of_years[year]:
                if profession.lower() in vacancy.name.lower():
                    count += 1
                    sum += vacancy.salary.get_salary_in_rur().get_mean_salary()
            mean_salaries_of_years[year] = 0 if count == 0 else math.floor(sum / count)
        return mean_salaries_of_years

    def get_num_of_vacancies_per_year(self, data, profession=''):
        """
        Возвращает словарь, состоящий из пар 'год - количество вакансий' для всего DataSet или для указанной профессии

        Args:
            data (DataSet): Объект DataSet, содержащий список вакансий
            profession (str): Название профессии, для которой нужно составить статистику. При отсутствии этого аргумента возвращается статистика по всем вакансиям

        Return:
            dict[int, int]: словарь, состоящий из пар 'год - количество вакансий'
        """
        vacancies = data.vacancies_objects
        vacancies_of_years = Utils.split_list(vacancies, lambda x: Utils.get_year(x.published_at))
        vacancies_num_of_years = {}
        for year in vacancies_of_years:
            count = 0
            for vacancy in vacancies_of_years[year]:
                if profession.lower() in vacancy.name.lower():
                    count += 1
            vacancies_num_of_years[year] = count
        return vacancies_num_of_years

    def get_salary_levels_of_areas(self, data):
        """
        Возвращает словарь, состоящий из пар 'регион - средняя зарплата' для всего DataSet или для указанной профессии

        Args:
            data (DataSet): Объект DataSet, содержащий список вакансий

        Return:
            dict[str, int]: словарь, состоящий из пар 'регион - средняя зарплата'
        """
        vacancies = data.vacancies_objects
        vacancies_of_areas = Utils.split_list(vacancies, lambda x: x.area_name)
        salary_levels_of_areas = {}
        for area in vacancies_of_areas:
            if len(vacancies_of_areas[area]) >= math.floor(self.total_vacancies * 0.01):
                sum = 0
                for vacancy in vacancies_of_areas[area]:
                    sum += vacancy.salary.get_salary_in_rur().get_mean_salary()
                salary_levels_of_areas[area] = math.floor(sum / len(vacancies_of_areas[area]))
        return dict(sorted(salary_levels_of_areas.items(), key=lambda x: x[1], reverse=True))

    def get_vacancy_fractions_of_areas(self, data):
        """
        Возвращает словарь, состоящий из пар 'регион - доля вакансий' для всего DataSet или для указанной профессии

        Args:
            data (DataSet): Объект DataSet, содержащий список вакансий

        Return:
            dict[str, float]: словарь, состоящий из пар 'регион - доля вакансий'
        """
        vacancies = data.vacancies_objects
        vacancies_of_areas = Utils.split_list(vacancies, lambda x: x.area_name)
        fractions_for_areas = {}
        for area in vacancies_of_areas:
            if len(vacancies_of_areas[area]) >= math.floor(self.total_vacancies * 0.01):
                fractions_for_areas[area] = float('{:.4f}'.format(len(vacancies_of_areas[area]) / self.total_vacancies))
        return dict(sorted(fractions_for_areas.items(), key=lambda x: x[1], reverse=True))

    def print_full_stats(self):
        """Выводит в консоль статистические данные"""
        print(f'Динамика уровня зарплат по годам: {self.year_salary_dynamics}')
        print(f'Динамика количества вакансий по годам: {self.num_of_vacancies_per_year}')
        print(f'Динамика уровня зарплат по годам для выбранной профессии: {self.year_salary_dynamics_for_prof}')
        print(f'Динамика количества вакансий по годам для выбранной профессии: {self.num_of_vacancies_per_year_for_prof}')
        print(f'Уровень зарплат по городам (в порядке убывания): {Utils.get_first_dict_elements(self.salary_levels_of_areas, 10)}')
        print(f'Доля вакансий по городам (в порядке убывания): {Utils.get_first_dict_elements(self.vacancy_fractions_of_areas, 10)}')


class Report:
    """
    Класс, отвечающий за оформление статистических данных

    Attributes:
        profession (str): Профессия, по которой требуется статистика
        year_salary_dynamics (dict[int, int]): Динамика уровня зарплат по годам
        num_of_vacancies_per_year (dict[int, int]): Динамика количества вакансий по годам
        year_salary_dynamics_for_prof (dict[int, int]): Динамика уровня зарплат по годам для выбранной профессии
        num_of_vacancies_per_year_for_prof (dict[int, int]): Динамика количества вакансий по годам для выбранной профессии
        salary_levels_of_areas (dict[str, int]): Уровень зарплат по городам (в порядке убывания)
        vacancy_fractions_of_areas (dict[str, float]): Доля вакансий по городам (в порядке убывания)
    """
    bold_font = Font(name='Cambria', size=11, bold=True)
    normal_font = Font(name='Calibri', size=11, bold=False)
    black_border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000'), )

    def __init__(
            self,
            profession,
            year_salary_dynamics,
            num_of_vacancies_per_year,
            year_salary_dynamics_for_prof,
            num_of_vacancies_per_year_for_prof,
            salary_levels_of_areas,
            vacancy_fractions_of_areas):
        """
        Инициализация объекта класса

        Args:
            profession (str): Профессия, по которой требуется статистика
            year_salary_dynamics (dict[int, int]): Динамика уровня зарплат по годам
            num_of_vacancies_per_year (dict[int, int]): Динамика количества вакансий по годам
            year_salary_dynamics_for_prof (dict[int, int]): Динамика уровня зарплат по годам для выбранной профессии
            num_of_vacancies_per_year_for_prof (dict[int, int]): Динамика количества вакансий по годам для выбранной профессии
            salary_levels_of_areas (dict[str, int]): Уровень зарплат по городам (в порядке убывания)
            vacancy_fractions_of_areas (dict[str, float]): Доля вакансий по городам (в порядке убывания)
        """
        self.profession = profession
        self.year_salary_dynamics = year_salary_dynamics
        self.num_of_vacancies_per_year = num_of_vacancies_per_year
        self.year_salary_dynamics_for_prof = year_salary_dynamics_for_prof
        self.num_of_vacancies_per_year_for_prof = num_of_vacancies_per_year_for_prof
        self.salary_levels_of_areas = salary_levels_of_areas
        self.vacancy_fractions_of_areas = vacancy_fractions_of_areas

    @classmethod
    def generate_excel(
            cls,
            profession,
            year_salary_dynamics,
            num_of_vacancies_per_year,
            year_salary_dynamics_for_prof,
            num_of_vacancies_per_year_for_prof,
            salary_levels_of_areas,
            vacancy_fractions_of_areas
    ):
        """
        Создаёт Excel-таблицу статистики

        Args:
            profession (str): Профессия, по которой требуется статистика
            year_salary_dynamics (dict[int, int]): Динамика уровня зарплат по годам
            num_of_vacancies_per_year (dict[int, int]): Динамика количества вакансий по годам
            year_salary_dynamics_for_prof (dict[int, int]): Динамика уровня зарплат по годам для выбранной профессии
            num_of_vacancies_per_year_for_prof (dict[int, int]): Динамика количества вакансий по годам для выбранной профессии
            salary_levels_of_areas (dict[str, int]): Уровень зарплат по городам (в порядке убывания)
            vacancy_fractions_of_areas (dict[str, float]): Доля вакансий по городам (в порядке убывания)
        """
        def apply_label_style(cell):
            """
            Добавляет нужный стиль ячейке заголовков

            Args:
                cell: Ячейка таблицы
            """
            cell.font = Report.bold_font
            cell.border = Report.black_border

        def apply_data_style(cell):

            """
            Добавляет нужный стиль ячейке значений

            Args:
                cell: Ячейка таблицы
            """
            cell.font = Report.normal_font
            cell.border = Report.black_border

        def fill_column(sheet, data, column_number, label=''):
            """
            Заполняет столбец таблицы значениями

            Args:
                sheet: Excel-лист
                data (list[Any]): Список, значениями элементов которого заполняется столбец
                column_number (int): Номер столбца
                label (str): Заголовок столбца. При непередаче этого авргумента, заголовок не добавляется
            """
            start_from = 1
            if label != '':
                start_from = 2
                cell = sheet.cell(row=1, column=column_number)
                cell.value = label
                apply_label_style(cell)
            for i in range(len(data)):
                cell = sheet.cell(row=i + start_from, column=column_number)
                cell.value = data[i]
                apply_data_style(cell)

        def connect_values_column(sheet, dictionary, key_column, target_column, label=''):
            """
            Добавляет в таблицу столбец значений аргумента-словаря в соответствие столбцу ключей

            Args:
                sheet: Excel-лист
                dictionary (dict[]): Словарь, ключи которого соответствуют значениям key_column
                key_column (int): Номер столбца ключей
                target_column (int): Номер заполняемого значениями столбца
                label (str): Заголовок столбца значений
            """
            start_from = 1
            if label != '':
                start_from = 2
                cell = sheet.cell(row=1, column=target_column)
                cell.value = label
                apply_label_style(cell)
            for i in range(len(dictionary)):
                key_cell = sheet.cell(row=i + start_from, column=key_column)
                value_cell = sheet.cell(row=i + start_from, column=target_column)
                value_cell.value = dictionary[key_cell.value]
                apply_data_style(value_cell)

        def adjust_col_width(sheet):
            """
            Определяет ширину ячеек в соответствии с содержимым

            Args:
                sheet: Excel-лист
            """
            widths = {}
            for row in sheet.rows:
                for cell in row:
                    if cell.value:
                        widths[cell.column_letter] = max((widths.get(cell.column_letter, 0), len(str(cell.value))))
            for col, value in widths.items():
                sheet.column_dimensions[col].width = value + 3

        book = Workbook()
        year_stats = book.active
        year_stats.title = 'Статистика по годам'
        area_stats = book.create_sheet('Статистика по городам')

        for col in area_stats.iter_cols(min_col=5, max_col=5, max_row=11):
            for cell in col:
                cell.number_format = '0.00%'

        is_prof_needed = profession != ''
        years = list(year_salary_dynamics.keys())

        fill_column(year_stats, years, 1, 'Год')
        connect_values_column(year_stats, year_salary_dynamics, 1, 2, 'Средняя зарплата')
        if is_prof_needed:
            connect_values_column(year_stats, year_salary_dynamics_for_prof, 1, 3, f'Средняя зарплата - {profession}')
        connect_values_column(year_stats, num_of_vacancies_per_year, 1, 4 if is_prof_needed else 3, 'Количество вакансий')
        if is_prof_needed:
            connect_values_column(year_stats, num_of_vacancies_per_year_for_prof, 1, 5, f'Количество вакансий - {profession}')

        fill_column(area_stats, list(salary_levels_of_areas.keys())[:10], 1, 'Город')
        connect_values_column(area_stats, Utils.get_first_dict_elements(salary_levels_of_areas, 10), 1, 2, 'Уровень зарплат')

        fill_column(area_stats, list(vacancy_fractions_of_areas.keys())[:10], 4, 'Город')
        connect_values_column(area_stats, Utils.get_first_dict_elements(vacancy_fractions_of_areas, 10), 4, 5, 'Доля вакансий')

        adjust_col_width(year_stats)
        adjust_col_width(area_stats)

        book.save('report.xlsx')

    @classmethod
    def generate_image(
            cls,
            profession,
            year_salary_dynamics,
            num_of_vacancies_per_year,
            year_salary_dynamics_for_prof,
            num_of_vacancies_per_year_for_prof,
            salary_levels_of_areas,
            vacancy_fractions_of_areas
    ):
        """
        Создаёт png-файл с графиками статистики

        Args:
            profession (str): Профессия, по которой требуется статистика
            year_salary_dynamics (dict[int, int]): Динамика уровня зарплат по годам
            num_of_vacancies_per_year (dict[int, int]): Динамика количества вакансий по годам
            year_salary_dynamics_for_prof (dict[int, int]): Динамика уровня зарплат по годам для выбранной профессии
            num_of_vacancies_per_year_for_prof (dict[int, int]): Динамика количества вакансий по годам для выбранной профессии
            salary_levels_of_areas (dict[str, int]): Уровень зарплат по городам (в порядке убывания)
            vacancy_fractions_of_areas (dict[str, float]): Доля вакансий по городам (в порядке убывания)
        """
        fig = pyplot.figure()

        year_salary_graph = fig.add_subplot(2, 2, 1)
        year_count_graph = fig.add_subplot(2, 2, 2)
        city_salary_graph = fig.add_subplot(2, 2, 3)
        city_frac_graph = fig.add_subplot(2, 2, 4)

        col_width = 0.4
        x_axis = np.arange(len(year_salary_dynamics.keys()))
        year_salary_graph.bar(x_axis - col_width / 2, year_salary_dynamics.values(), width=col_width, label='средняя з/п')
        year_salary_graph.bar(x_axis + col_width / 2, year_salary_dynamics_for_prof.values(), width=col_width, label=f'з/п {profession}')
        year_salary_graph.set(title='Уровень зарплат по годам', xticks=x_axis)
        year_salary_graph.set_xticklabels(labels=year_salary_dynamics.keys(), rotation='vertical', va='top', ha='center')
        year_salary_graph.tick_params(axis='both', labelsize=8)
        year_salary_graph.grid(True, axis='y')
        year_salary_graph.legend(fontsize=8)

        year_count_graph.set_title('Количество вакансий по годам')
        year_count_graph.bar(x_axis - col_width / 2, num_of_vacancies_per_year.values(), width=col_width, label='Количество вакансий')
        year_count_graph.bar(x_axis + col_width / 2, num_of_vacancies_per_year_for_prof.values(), width=col_width, label=f'Количество вакансий\n{profession}')
        year_count_graph.set_xticks(ticks=x_axis, labels=num_of_vacancies_per_year.keys())
        year_count_graph.set_xticklabels(labels=num_of_vacancies_per_year.keys(), rotation='vertical', va='top', ha='center')
        year_count_graph.tick_params(axis='both', labelsize=8)
        year_count_graph.grid(True, axis='y')
        year_count_graph.legend(fontsize=8)

        result_dic = Utils.get_first_dict_elements(salary_levels_of_areas, 10)
        city_salary_graph.set_title("Уровень зарплат по городам")
        city_salary_graph.invert_yaxis()
        areas = list(result_dic.keys())
        areas = [area.replace(' ', '\n').replace('-', '-\n') for area in areas]
        city_salary_graph.barh(areas, list(result_dic.values()))
        city_salary_graph.tick_params(axis='both', labelsize=8)
        city_salary_graph.set_yticklabels(areas, fontsize=6, va='center', ha='right')
        city_salary_graph.grid(True, axis='x')

        result_dic = Utils.get_first_dict_elements(vacancy_fractions_of_areas, 10)
        others = 1 - sum((list(result_dic.values())))
        result_dic.update({'Другие': others})
        labels = list(result_dic.keys())
        city_frac_graph.pie(list(result_dic.values()), labels=labels, textprops={'fontsize': 6})
        city_frac_graph.axis('scaled')
        city_frac_graph.set_title("Доля вакансий по городам")
        pyplot.tight_layout()
        pyplot.savefig('graph.png', dpi=300)

    @classmethod
    def generate_pdf(
            cls,
            profession,
            year_salary_dynamics,
            num_of_vacancies_per_year,
            year_salary_dynamics_for_prof,
            num_of_vacancies_per_year_for_prof,
            salary_levels_of_areas,
            vacancy_fractions_of_areas,
            graph_path
    ):
        """
        Создаёт pdf-файл, содержащий статистические графики и таблицы

        Args:
            profession (str): Профессия, по которой требуется статистика
            year_salary_dynamics (dict[int, int]): Динамика уровня зарплат по годам
            num_of_vacancies_per_year (dict[int, int]): Динамика количества вакансий по годам
            year_salary_dynamics_for_prof (dict[int, int]): Динамика уровня зарплат по годам для выбранной профессии
            num_of_vacancies_per_year_for_prof (dict[int, int]): Динамика количества вакансий по годам для выбранной профессии
            salary_levels_of_areas (dict[str, int]): Уровень зарплат по городам (в порядке убывания)
            vacancy_fractions_of_areas (dict[str, float]): Доля вакансий по городам (в порядке убывания)
            graph_path (str): путь к графику
        """

        salary_levels_of_areas = Utils.get_first_dict_elements(salary_levels_of_areas, 10)
        vacancy_fractions_of_areas = Utils.get_first_dict_elements(vacancy_fractions_of_areas, 10)

        vacancy_fractions_of_areas = vacancy_fractions_of_areas.items()
        vacancy_fractions_of_areas = {key: str(f'{value * 100:,.2f}%').replace('.', ',') for (key, value) in vacancy_fractions_of_areas}

        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")

        image_header = f'Аналитика по зарплатам и городам для профессии {profession}'
        years_table_header = 'Статистика по годам'
        areas_table_header = 'Статистика по городам'
        years_table_labels = ['Год',
                              'Средняя зарплата',
                              f'Средняя зарплата - {profession}',
                              'Количество вакансий',
                              f'Количество вакансий - {profession}']

        area_salaries_table_labels = ['Город', 'Уровень зарплат']
        area_fracs_table_labels = ['Город', 'Доля вакансий']

        years = list(year_salary_dynamics.keys())
        areas_for_salaries = list(salary_levels_of_areas.keys())
        areas_for_fracs = list(vacancy_fractions_of_areas.keys())

        pdf_template = template.render({
                                        'image_header': image_header,
                                        'profession': profession,
                                        'years': years,
                                        'years_table_header': years_table_header,
                                        'areas_table_header': areas_table_header,
                                        'years_table_labels': years_table_labels,
                                        'area_salaries_table_labels': area_salaries_table_labels,
                                        'area_fracs_table_labels': area_fracs_table_labels,
                                        'areas_for_salaries': areas_for_salaries,
                                        'areas_for_fracs': areas_for_fracs,
                                        'year_salary_dynamics': year_salary_dynamics,
                                        'num_of_vacancies_per_year': num_of_vacancies_per_year,
                                        'year_salary_dynamics_for_prof': year_salary_dynamics_for_prof,
                                        'num_of_vacancies_per_year_for_prof': num_of_vacancies_per_year_for_prof,
                                        'salary_levels_of_areas': salary_levels_of_areas,
                                        'vacancy_fractions_of_areas': vacancy_fractions_of_areas,
                                        'graph': graph_path})
        config = pdfkit.configuration(wkhtmltopdf=r'D:\wkhtmltox\bin\wkhtmltopdf.exe')

        options = {'enable-local-file-access': True}
        pdfkit.from_string(pdf_template, 'report.pdf', options=options, configuration=config)

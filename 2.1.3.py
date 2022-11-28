import csv
import re
from datetime import datetime
from itertools import islice
import os

from prettytable import PrettyTable
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
import numpy as np
import matplotlib.pyplot as plt
import math
from jinja2 import Environment, FileSystemLoader
import pdfkit


class Dicts:
    dic_naming = {
        'name': 'Название',
        'description': 'Описание',
        'key_skills': 'Навыки',
        'experience_id': 'Опыт работы',
        'premium': 'Премиум-вакансия',
        'employer_name': 'Компания',
        'salary_from': 'Нижняя граница вилки оклада',
        'salary_to': 'Верхняя граница вилки оклада',
        'salary_gross': 'Оклад указан до вычета налогов',
        'salary_currency': 'Идентификатор валюты оклада',
        'area_name': 'Название региона',
        'published_at': 'Дата и время публикации вакансии',
        'FALSE': 'Нет',
        'TRUE': 'Да',
        'False': 'Нет',
        'True': 'Да',
        'noExperience': 'Нет опыта',
        'between1And3': 'От 1 года до 3 лет',
        'between3And6': 'От 3 до 6 лет',
        'moreThan6': 'Более 6 лет',
        'AZN': 'Манаты',
        'BYR': 'Белорусские рубли',
        'EUR': 'Евро',
        'GEL': 'Грузинский лари',
        'KGS': 'Киргизский сом',
        'KZT': 'Тенге',
        'RUR': 'Рубли',
        'UAH': 'Гривны',
        'USD': 'Доллары',
        'UZS': 'Узбекский сум'
    }

    currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

    experience_in_numbers = {
        'noExperience': 0,
        'between1And3': 3,
        'between3And6': 6,
        'moreThan6': 7,
    }


class Utils:
    @staticmethod
    def cut_string(string):
        if string is None:
            return None
        if len(string) > 100:
            return string[0: 100] + '...'
        else:
            return string

    @staticmethod
    def format_num_string(num):
        num = Utils.cut_frac(num)
        num = num[::-1]
        return ' '.join(num[i:i + 3] for i in range(0, len(num), 3))[::-1]

    @staticmethod
    def format_date(string):
        return {'output': f'{string[8:10]}.{string[5:7]}.{string[0:4]}',
                'time': datetime.strptime(string[:-5], "%Y-%m-%dT%H:%M:%S")}

    @staticmethod
    def get_split_count(string, sep):
        if string == '':
            return 0
        else:
            return len(string.split(sep))

    @staticmethod
    def get_year(string):
        return datetime.strptime(string[:-5], "%Y-%m-%dT%H:%M:%S").year

    @staticmethod
    def format_string(input_string):
        result = re.sub(r'<[^<>]*>', '', input_string)
        result = re.sub(r'\n', '!crutch!!', result)
        result = str.strip(re.sub(r'\s+', ' ', result))
        return result

    @staticmethod
    def cut_frac(string):
        if '.' in string:
            return string[:string.find('.')]
        return string

    @staticmethod
    def split_list(original, key_getter):
        groups = {}
        keys = []
        for element in original:
            key = key_getter(element)
            if key not in keys:
                keys.append(key)
                groups[key] = list()
            groups[key].append(element)
        return groups

    @staticmethod
    def get_first_dict_elements(dictionary, n):
        result = {}
        if n < 0:
            raise ValueError('index out of range')
        for key in list(islice(dictionary, n)):
            result[key] = dictionary[key]
        return result


class Salary:
    def __init__(self, salary_from, salary_to, salary_gross, currency):
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_gross = salary_gross
        self.salary_currency = currency


class Vacancy:
    def __init__(self,
                 name,
                 description,
                 key_skills,
                 experience_id,
                 premium,
                 employer_name,
                 salary_from,
                 salary_to,
                 salary_gross,
                 salary_currency,
                 area_name,
                 published_at):
        self.name = name
        self.description = description
        self.key_skills = key_skills
        self.experience_id = experience_id
        self.premium = premium
        self.employer_name = employer_name
        self.salary = Salary(salary_from, salary_to, salary_gross, salary_currency)
        self.area_name = area_name
        self.published_at = published_at


class DataSet:
    sorting = {
        'Название': lambda x: x.name,
        'Описание': lambda x: x.description,
        'Компания': lambda x: x.employer_name,
        'Навыки': lambda x: len(x.key_skills),
        'Опыт работы': lambda x: Dicts.experience_in_numbers[x.experience_id],
        'Премиум-вакансия': lambda x: x.premium,
        'Оклад': lambda x: math.floor(
            (int(Utils.cut_frac(x.salary.salary_from)) + int(Utils.cut_frac(x.salary.salary_to))) / 2) *
                           Dicts.currency_to_rub[x.salary.salary_currency],
        'Название региона': lambda x: x.area_name,
        'Дата публикации вакансии': lambda x: Utils.format_date(x.published_at)['time'],
        'Идентификатор валюты оклада': lambda x: x.salary.salary_currency,
        'Оклад указан до вычета налогов': lambda x: x.salary.salary_gross
    }

    def __init__(self, file_name):
        self.file_name = file_name
        self.vacancies_objects = []
        vacancies = []
        labels = []
        with open(file_name, encoding='utf-8-sig') as data:
            reader = csv.reader(data, delimiter=',')
            checker = 0
            for row in reader:
                if checker == 0:
                    checker += 1
                    labels = row
                else:
                    vacancies.append(row)
        cleaned_rows = []
        for row in vacancies:
            if all(row) and len(labels) == len(row):
                temp = []
                for i in range(len(row)):
                    temp.append(Utils.format_string(row[i]))
                cleaned_rows.append(temp)
        for row in cleaned_rows:
            vacancy = {}
            for i in range(len(labels)):
                vacancy[labels[i]] = row[i]

            def check_presence(vacancy, label):
                return None if label not in vacancy.keys() else vacancy[label]

            self.vacancies_objects.append(Vacancy(check_presence(vacancy, 'name'),
                                                  check_presence(vacancy, 'description'),
                                                  None if 'key_skills' not in vacancy.keys() else vacancy['key_skills'].split('!crutch!!'),
                                                  check_presence(vacancy, 'experience_id'),
                                                  check_presence(vacancy, 'premium'),
                                                  check_presence(vacancy, 'employer_name'),
                                                  check_presence(vacancy, 'salary_from'),
                                                  check_presence(vacancy, 'salary_to'),
                                                  check_presence(vacancy, 'salary_gross'),
                                                  check_presence(vacancy, 'salary_currency'),
                                                  check_presence(vacancy, 'area_name'),
                                                  check_presence(vacancy, 'published_at')))

    def length(self):
        return len(self.vacancies_objects)

    def sort(self, sorting_criteria, is_reversed):
        is_reversed = is_reversed == 'Да'
        self.vacancies_objects.sort(key=self.sorting[sorting_criteria], reverse=is_reversed)

    @staticmethod
    def format_filter_criteria(filter_criteria):
        filter_criteria = filter_criteria.split(': ')
        return {
            'label': filter_criteria[0] if filter_criteria[0] not in Dicts.dic_naming.keys() else Dicts.dic_naming[
                filter_criteria[0]],
            'content': '' if filter_criteria[0] == '' else (
                filter_criteria[1] if filter_criteria[1] not in Dicts.dic_naming.keys() else Dicts.dic_naming[
                    filter_criteria[1]])
        }

    def filter(self, filter_criteria):
        filter_criteria = self.format_filter_criteria(filter_criteria)
        filtering = {
            '': lambda x: True,
            'Название': lambda x: x.name == filter_criteria['content'],
            'Описание': lambda x: x.description == filter_criteria['content'],
            'Компания': lambda x: x.employer_name == filter_criteria['content'],
            'Навыки': lambda x: set(filter_criteria['content'].split(', ')).issubset(x.key_skills),
            'Опыт работы': lambda x: Dicts.dic_naming[x.experience_id] == filter_criteria['content'],
            'Премиум-вакансия': lambda x: Dicts.dic_naming[x.premium] == filter_criteria['content'],
            'Оклад': lambda x: int(Utils.cut_frac(x.salary.salary_from)) <= int(filter_criteria['content']) <= int(
                Utils.cut_frac(x.salary.salary_to)),
            'Название региона': lambda x: x.area_name == filter_criteria['content'],
            'Дата публикации вакансии': lambda x: Utils.format_date(x.published_at)['output'] == filter_criteria[
                'content'],
            'Идентификатор валюты оклада': lambda x: Dicts.dic_naming[x.salary.salary_currency] == filter_criteria[
                'content'],
            'Оклад указан до вычета налогов': lambda x: Dicts.dic_naming[x.salary.salary_gross] == filter_criteria[
                'content']
        }
        self.vacancies_objects = list(filter(filtering[filter_criteria['label']], self.vacancies_objects))


class TablePrinter:
    possible_criteria = ['Название', 'Описание', 'Навыки', 'Опыт работы', 'Премиум-вакансия',
                         'Компания', 'Оклад', 'Название региона', 'Дата публикации вакансии',
                         'Идентификатор валюты оклада', 'Оклад указан до вычета налогов']

    def __init__(self, data_set):
        self.data = data_set
        self.filter_criteria = input('Введите параметр фильтрации: ')
        self.sorting_criteria = input('Введите параметр сортировки: ')
        self.sort_reversed = input('Обратный порядок сортировки (Да / Нет): ')
        self.from_to = input('Введите диапазон вывода: ')
        self.fields = input('Введите требуемые столбцы: ')

    def print_table(self, dictionary):
        if self.filter_criteria != '' and not ':' in self.filter_criteria:
            print('Формат ввода некорректен')
            return
        if self.filter_criteria.split(': ')[0] != '' and not self.filter_criteria.split(': ')[
                                                                 0] in self.possible_criteria:
            print('Параметр поиска некорректен')
            return
        if self.sorting_criteria not in self.possible_criteria and self.sorting_criteria != '':
            print('Параметр сортировки некорректен')
            return
        if self.sort_reversed not in ['Да', 'Нет', '']:
            print('Порядок сортировки задан некорректно')
            return
        if os.stat(self.data.file_name).st_size == 0:
            print("Пустой файл")
            return

        if self.data.length() == 0:
            print('Нет данных')
            return
        if self.sorting_criteria != '':
            self.data.sort(self.sorting_criteria, self.sort_reversed)

        self.data.filter(self.filter_criteria)
        if len(self.data.vacancies_objects) == 0:
            print('Ничего не найдено')
            return
        labels = ['№', 'Название', 'Описание', 'Навыки', 'Опыт работы', 'Премиум-вакансия',
                  'Компания', 'Оклад', 'Название региона', 'Дата публикации вакансии']
        table = PrettyTable(field_names=labels)

        table.hrules = 1
        table.max_width = 20
        table.align = 'l'

        for i in range(self.data.length()):
            vacancy = self.data.vacancies_objects[i]
            salary = vacancy.salary
            table.add_row([
                i + 1,
                vacancy.name,
                Utils.cut_string(vacancy.description),
                Utils.cut_string('\n'.join(vacancy.key_skills)),
                dictionary[vacancy.experience_id],
                dictionary[vacancy.premium],
                vacancy.employer_name,
                f'{Utils.format_num_string(salary.salary_from)} - {Utils.format_num_string(salary.salary_to)} ({dictionary[salary.salary_currency]}) ' + (
                    '(Без вычета налогов)' if (
                                salary.salary_gross == 'TRUE' or salary.salary_gross == 'true' or salary.salary_gross == 'True') else '(С вычетом налогов)'),
                vacancy.area_name,
                Utils.format_date(vacancy.published_at)['output']
            ])

        border_variant = Utils.get_split_count(self.from_to, ' ')
        from_to = self.from_to.split(' ')
        fields = self.fields

        cut_borders_variants = {
            0: lambda: (0, self.data.length()),
            1: lambda: ((int(from_to[0]) - 1), self.data.length()),
            2: lambda: (int(from_to[0]) - 1, int(from_to[1]) - 1)
        }
        start = cut_borders_variants[border_variant]()[0]
        end = cut_borders_variants[border_variant]()[1]
        if fields != '':
            fields = fields.split(', ')
            fields.insert(0, '№')
            print(table.get_string(start=start, end=end, fields=fields))
        else:
            print(table.get_string(start=start, end=end))


class Stats:
    bold_font = Font(name='Cambria', size=11, bold=True)
    normal_font = Font(name='Calibri', size=11, bold=False)
    black_border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000'),)

    def __init__(self, data):
        self.profession = input('Введите название профессии: ')
        self.data = data
        self.total_vacancies = self.data.length()
        self.year_salary_dynamics = self.get_year_salary_dynamics()
        self.num_of_vacancies_per_year = self.get_num_of_vacancies_per_year()
        self.year_salary_dynamics_for_prof = self.get_year_salary_dynamics(self.profession)
        self.num_of_vacancies_per_year_for_prof = self.get_num_of_vacancies_per_year(self.profession)
        self.salary_levels_of_areas = self.get_salary_levels_of_areas()
        self.vacancy_fractions_of_areas = self.get_vacancy_fractions_of_areas()

    def get_year_salary_dynamics(self, profession=''):
        vacancies = self.data.vacancies_objects
        vacancies_of_years = Utils.split_list(vacancies, lambda x: Utils.get_year(x.published_at))
        mean_salaries_of_years = {}
        for year in vacancies_of_years:
            sum = 0
            count = 0
            for vacancy in vacancies_of_years[year]:
                if profession.lower() in vacancy.name.lower():
                    count += 1
                    sum += math.floor((int(Utils.cut_frac(vacancy.salary.salary_from)) + int(Utils.cut_frac(vacancy.salary.salary_to))) / 2) * Dicts.currency_to_rub[vacancy.salary.salary_currency]
            mean_salaries_of_years[year] = 0 if count == 0 else math.floor(sum / count)
        return mean_salaries_of_years

    def get_num_of_vacancies_per_year(self, profession=''):
        vacancies = self.data.vacancies_objects
        vacancies_of_years = Utils.split_list(vacancies, lambda x: Utils.get_year(x.published_at))
        vacancies_num_of_years = {}
        for year in vacancies_of_years:
            count = 0
            for vacancy in vacancies_of_years[year]:
                if profession.lower() in vacancy.name.lower():
                    count += 1
            vacancies_num_of_years[year] = count
        return vacancies_num_of_years

    def get_salary_levels_of_areas(self):
        vacancies = self.data.vacancies_objects
        vacancies_of_areas = Utils.split_list(vacancies, lambda x: x.area_name)
        salary_levels_of_areas = {}
        for area in vacancies_of_areas:
            if len(vacancies_of_areas[area]) >= math.floor(self.total_vacancies * 0.01):
                sum = 0
                for vacancy in vacancies_of_areas[area]:
                    sum += math.floor((int(Utils.cut_frac(vacancy.salary.salary_from)) + int(Utils.cut_frac(vacancy.salary.salary_to))) / 2) * Dicts.currency_to_rub[vacancy.salary.salary_currency]
                salary_levels_of_areas[area] = math.floor(sum / len(vacancies_of_areas[area]))
        return dict(sorted(salary_levels_of_areas.items(), key=lambda x: x[1], reverse=True))

    def get_vacancy_fractions_of_areas(self):
        vacancies = self.data.vacancies_objects
        vacancies_of_areas = Utils.split_list(vacancies, lambda x: x.area_name)
        fractions_for_areas = {}
        for area in vacancies_of_areas:
            if len(vacancies_of_areas[area]) >= math.floor(self.total_vacancies * 0.01):
                fractions_for_areas[area] = float('{:.4f}'.format(len(vacancies_of_areas[area]) / self.total_vacancies))
        return dict(sorted(fractions_for_areas.items(), key=lambda x: x[1], reverse=True))

    def print_full_stats(self):
        print(f'Динамика уровня зарплат по годам: {self.year_salary_dynamics}')
        print(f'Динамика количества вакансий по годам: {self.num_of_vacancies_per_year}')
        print(f'Динамика уровня зарплат по годам для выбранной профессии: {self.year_salary_dynamics_for_prof}')
        print(f'Динамика количества вакансий по годам для выбранной профессии: {self.num_of_vacancies_per_year_for_prof}')
        print(f'Уровень зарплат по городам (в порядке убывания): {Utils.get_first_dict_elements(self.salary_levels_of_areas, 10)}')
        print(f'Доля вакансий по городам (в порядке убывания): {Utils.get_first_dict_elements(self.vacancy_fractions_of_areas, 10)}')


class Report:
    def __init__(
            self,
            profession,
            year_salary_dynamics,
            num_of_vacancies_per_year,
            year_salary_dynamics_for_prof,
            num_of_vacancies_per_year_for_prof,
            salary_levels_of_areas,
            vacancy_fractions_of_areas):
        self.profession = profession
        self.year_salary_dynamics = year_salary_dynamics
        self.num_of_vacancies_per_year = num_of_vacancies_per_year
        self.year_salary_dynamics_for_prof = year_salary_dynamics_for_prof
        self.num_of_vacancies_per_year_for_prof = num_of_vacancies_per_year_for_prof
        self.salary_levels_of_areas = salary_levels_of_areas
        self.vacancy_fractions_of_areas = vacancy_fractions_of_areas

    @classmethod
    def generate_excel(
            cls,
            profession,
            year_salary_dynamics,
            num_of_vacancies_per_year,
            year_salary_dynamics_for_prof,
            num_of_vacancies_per_year_for_prof,
            salary_levels_of_areas,
            vacancy_fractions_of_areas
    ):

        def apply_label_style(cell):
            cell.font = Stats.bold_font
            cell.border = Stats.black_border

        def apply_data_style(cell):
            cell.font = Stats.normal_font
            cell.border = Stats.black_border

        def fill_column(sheet, data, column_number, label=''):
            start_from = 1
            if label != '':
                start_from = 2
                cell = sheet.cell(row=1, column=column_number)
                cell.value = label
                apply_label_style(cell)
            for i in range(len(data)):
                cell = sheet.cell(row=i + start_from, column=column_number)
                cell.value = data[i]
                apply_data_style(cell)

        def connect_values_column(sheet, dictionary, key_column, target_column, label=''):
            start_from = 1
            if label != '':
                start_from = 2
                cell = sheet.cell(row=1, column=target_column)
                cell.value = label
                apply_label_style(cell)
            for i in range(len(dictionary)):
                key_cell = sheet.cell(row=i + start_from, column=key_column)
                value_cell = sheet.cell(row=i + start_from, column=target_column)
                value_cell.value = dictionary[key_cell.value]
                apply_data_style(value_cell)

        def adjust_col_width(sheet):
            widths = {}
            for row in sheet.rows:
                for cell in row:
                    if cell.value:
                        widths[cell.column_letter] = max((widths.get(cell.column_letter, 0), len(str(cell.value))))
            for col, value in widths.items():
                sheet.column_dimensions[col].width = value + 3

        book = Workbook()
        year_stats = book.active
        year_stats.title = 'Статистика по годам'
        area_stats = book.create_sheet('Статистика по городам')

        for col in area_stats.iter_cols(min_col=5, max_col=5, max_row=11):
            for cell in col:
                cell.number_format = '0.00%'

        is_prof_needed = profession != ''
        years = list(year_salary_dynamics.keys())

        fill_column(year_stats, years, 1, 'Год')
        connect_values_column(year_stats, year_salary_dynamics, 1, 2, 'Средняя зарплата')
        if is_prof_needed:
            connect_values_column(year_stats, year_salary_dynamics_for_prof, 1, 3, f'Средняя зарплата - {profession}')
        connect_values_column(year_stats, num_of_vacancies_per_year, 1, 4 if is_prof_needed else 3, 'Количество вакансий')
        if is_prof_needed:
            connect_values_column(year_stats, num_of_vacancies_per_year_for_prof, 1, 5, f'Количество вакансий - {profession}')

        fill_column(area_stats, list(salary_levels_of_areas.keys())[:10], 1, 'Город')
        connect_values_column(area_stats, Utils.get_first_dict_elements(salary_levels_of_areas, 10), 1, 2, 'Уровень зарплат')

        fill_column(area_stats, list(vacancy_fractions_of_areas.keys())[:10], 4, 'Город')
        connect_values_column(area_stats, Utils.get_first_dict_elements(vacancy_fractions_of_areas, 10), 4, 5, 'Доля вакансий')

        adjust_col_width(year_stats)
        adjust_col_width(area_stats)

        book.save('report.xlsx')

    @classmethod
    def generate_image(
            cls,
            profession,
            year_salary_dynamics,
            num_of_vacancies_per_year,
            year_salary_dynamics_for_prof,
            num_of_vacancies_per_year_for_prof,
            salary_levels_of_areas,
            vacancy_fractions_of_areas
    ):
        fig = plt.figure()

        year_salary_graph = fig.add_subplot(2, 2, 1)
        year_count_graph = fig.add_subplot(2, 2, 2)
        city_salary_graph = fig.add_subplot(2, 2, 3)
        city_frac_graph = fig.add_subplot(2, 2, 4)

        col_width = 0.4
        x_axis = np.arange(len(year_salary_dynamics.keys()))
        year_salary_graph.bar(x_axis - col_width / 2, year_salary_dynamics.values(), width=col_width, label='средняя з/п')
        year_salary_graph.bar(x_axis + col_width / 2, year_salary_dynamics_for_prof.values(), width=col_width, label=f'з/п {profession}')
        year_salary_graph.set(title='Уровень зарплат по годам', xticks=x_axis)
        year_salary_graph.set_xticklabels(labels=year_salary_dynamics.keys(), rotation='vertical', va='top', ha='center')
        year_salary_graph.tick_params(axis='both', labelsize=8)
        year_salary_graph.grid(True, axis='y')
        year_salary_graph.legend(fontsize=8)

        year_count_graph.set_title('Количество вакансий по годам')
        year_count_graph.bar(x_axis - col_width / 2, num_of_vacancies_per_year.values(), width=col_width, label='Количество вакансий')
        year_count_graph.bar(x_axis + col_width / 2, num_of_vacancies_per_year_for_prof.values(), width=col_width, label=f'Количество вакансий\n{profession}')
        year_count_graph.set_xticks(ticks=x_axis, labels=num_of_vacancies_per_year.keys())
        year_count_graph.set_xticklabels(labels=num_of_vacancies_per_year.keys(), rotation='vertical', va='top', ha='center')
        year_count_graph.tick_params(axis='both', labelsize=8)
        year_count_graph.grid(True, axis='y')
        year_count_graph.legend(fontsize=8)

        result_dic = Utils.get_first_dict_elements(salary_levels_of_areas, 10)
        city_salary_graph.set_title("Уровень зарплат по городам")
        city_salary_graph.invert_yaxis()
        areas = list(result_dic.keys())
        areas = [area.replace(' ', '\n').replace('-', '-\n') for area in areas]
        city_salary_graph.barh(areas, list(result_dic.values()))
        city_salary_graph.tick_params(axis='both', labelsize=8)
        city_salary_graph.set_yticklabels(areas, fontsize=6, va='center', ha='right')
        city_salary_graph.grid(True, axis='x')

        result_dic = Utils.get_first_dict_elements(vacancy_fractions_of_areas, 10)
        others = 1 - sum((list(result_dic.values())))
        result_dic.update({'Другие': others})
        labels = list(result_dic.keys())
        city_frac_graph.pie(list(result_dic.values()), labels=labels, textprops={'fontsize': 6})
        city_frac_graph.axis('scaled')
        city_frac_graph.set_title("Доля вакансий по городам")
        plt.tight_layout()
        plt.savefig('graph.png', dpi=300)

    @classmethod
    def generate_pdf(
            cls,
            profession,
            year_salary_dynamics,
            num_of_vacancies_per_year,
            year_salary_dynamics_for_prof,
            num_of_vacancies_per_year_for_prof,
            salary_levels_of_areas,
            vacancy_fractions_of_areas
    ):

        salary_levels_of_areas = Utils.get_first_dict_elements(salary_levels_of_areas, 10)
        vacancy_fractions_of_areas = Utils.get_first_dict_elements(vacancy_fractions_of_areas, 10)

        vacancy_fractions_of_areas = vacancy_fractions_of_areas.items()
        vacancy_fractions_of_areas = {key: str(f'{value * 100:,.2f}%').replace('.', ',') for (key, value) in vacancy_fractions_of_areas}

        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")

        graph_path = r"C:\Okladnikov\graph.png"
        image_header = f'Аналитика по зарплатам и городам для профессии {profession}'
        years_table_header = 'Статистика по годам'
        areas_table_header = 'Статистика по городам'
        years_table_labels = ['Год',
                              'Средняя зарплата',
                              f'Средняя зарплата - {profession}',
                              'Количество вакансий',
                              f'Количество вакансий - {profession}']

        area_salaries_table_labels = ['Город', 'Уровень зарплат']
        area_fracs_table_labels = ['Город', 'Доля вакансий']

        years = list(year_salary_dynamics.keys())
        areas_for_salaries = list(salary_levels_of_areas.keys())
        areas_for_fracs = list(vacancy_fractions_of_areas.keys())

        pdf_template = template.render({
                                        'image_header': image_header,
                                        'profession': profession,
                                        'years': years,
                                        'years_table_header': years_table_header,
                                        'areas_table_header': areas_table_header,
                                        'years_table_labels': years_table_labels,
                                        'area_salaries_table_labels': area_salaries_table_labels,
                                        'area_fracs_table_labels': area_fracs_table_labels,
                                        'areas_for_salaries': areas_for_salaries,
                                        'areas_for_fracs': areas_for_fracs,
                                        'year_salary_dynamics': year_salary_dynamics,
                                        'num_of_vacancies_per_year': num_of_vacancies_per_year,
                                        'year_salary_dynamics_for_prof': year_salary_dynamics_for_prof,
                                        'num_of_vacancies_per_year_for_prof': num_of_vacancies_per_year_for_prof,
                                        'salary_levels_of_areas': salary_levels_of_areas,
                                        'vacancy_fractions_of_areas': vacancy_fractions_of_areas,
                                        'graph': graph_path})
        config = pdfkit.configuration(wkhtmltopdf=r'D:\wkhtmltox\bin\wkhtmltopdf.exe')

        options = {'enable-local-file-access': True}
        pdfkit.from_string(pdf_template, 'report.pdf', options=options, configuration=config)


def print_vacancies_table(data):
    printer = TablePrinter(data)
    printer.print_table(Dicts.dic_naming)


def report_stats(data):
    stats = Stats(data)
    stats.print_full_stats()
    report = Report(stats.profession,
                    stats.year_salary_dynamics,
                    stats.num_of_vacancies_per_year,
                    stats.year_salary_dynamics_for_prof,
                    stats.num_of_vacancies_per_year_for_prof,
                    stats.salary_levels_of_areas,
                    stats.vacancy_fractions_of_areas)

    Report.generate_image(
        profession=report.profession,
        year_salary_dynamics=report.year_salary_dynamics,
        num_of_vacancies_per_year=report.num_of_vacancies_per_year,
        year_salary_dynamics_for_prof=report.year_salary_dynamics_for_prof,
        num_of_vacancies_per_year_for_prof=report.num_of_vacancies_per_year_for_prof,
        salary_levels_of_areas=report.salary_levels_of_areas,
        vacancy_fractions_of_areas=report.vacancy_fractions_of_areas
    )

    Report.generate_pdf(
        profession=report.profession,
        year_salary_dynamics=report.year_salary_dynamics,
        num_of_vacancies_per_year=report.num_of_vacancies_per_year,
        year_salary_dynamics_for_prof=report.year_salary_dynamics_for_prof,
        num_of_vacancies_per_year_for_prof=report.num_of_vacancies_per_year_for_prof,
        salary_levels_of_areas=report.salary_levels_of_areas,
        vacancy_fractions_of_areas=report.vacancy_fractions_of_areas
    )


commands = {'Вакансии': lambda data: print_vacancies_table(data),
            'Статистика': lambda data: report_stats(data)}

command = input('Введите команду: ')
if command not in list(commands.keys()):
    print('Неизвестная команда!')
else:
    commands[command](DataSet(input('Введите данные для печати: ')))
